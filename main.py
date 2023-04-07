# -*- coding: utf8 -*-
import os
import re
import time
import lxml
import shutil
import sys
import aiohttp
import asyncio
import aiofiles
import requests
from bs4 import BeautifulSoup
from fake_useragent import UserAgent
from openpyxl import load_workbook
from tqdm import tqdm
import datetime
from PIL import Image, ImageFile


class Parser:

    def __init__(self):
        ua = UserAgent()
        self.headers = {'user_agent': ua.random}
        self.token = ''
        self.secret_key = ''
        self.active_token = ''
        self.active_secret_key = ''
        self.base_url = 'https://berlinger-haus-shop.ru'
        self.article_numbers = []
        self.links_products = {}
        self.article_imgs = {}
        self.article_save_imgs = {}
        self.read_data_file = ''

    def open_token_file(self):
        try:
            with open('token.txt', 'r') as file:
                for i, line in enumerate(file):
                    if i == 0:
                        self.token = line.split('=')[1].strip().split(', ')
                    elif i == 1:
                        self.secret_key = line.split('=')[1].strip().split(', ')
        except Exception:
            print('Не удалось прочитать token или secret_key')
            raise IndexError

    def read_file(self):
        try:
            for file in os.listdir():
                if file[:6] == 'data1.':
                    self.read_data_file = file
        except Exception:
            print('Нет файла с именем data.')
            raise IndexError

    def get_article_number(self):
        try:
            wb = load_workbook(filename=self.read_data_file)
            sheets = wb.sheetnames
            ws = wb[sheets[0]]

            for row in ws.iter_cols(min_col=2, max_col=2, min_row=9):
                for cell in row:
                    if cell.value is None:
                        continue
                    # есть ли числа в строке
                    if re.search('\d+', cell.value.strip().split(' ')[0]):
                        self.article_numbers.append(cell.value.strip().split(' ')[0])

            self.article_numbers = list(dict.fromkeys(self.article_numbers))
        except Exception as exc:
            print(f'Ошибка {exc} в чтении табличного документа data.xlsx')
            with open('error.txt', 'a', encoding='utf-8') as file:

                file.write(f'{datetime.datetime.now().strftime("%d-%m-%y %H:%M")} '
                           f'Ошибка {exc} в чтении табличного документа data.xlsm, функция - get_article_number()\n')
            raise IndexError

    def get_link_prodicts(self):
        for art in self.article_numbers:

            sys.stdout.write("\r")
            sys.stdout.write(f'Обрабатываю - {art}')
            sys.stdout.flush()

            if len(art) == 1:
                continue
            response = requests.get(f'{self.base_url}/search?q={art[3:]}', headers=self.headers)
            soup = BeautifulSoup(response.text, features='lxml')

            product_not_found = soup.find('p', class_='warning')
            if bool(product_not_found) is False:

                if f'{self.base_url}/search?q={art[3:]}' == response.url:
                    found_links_some_product = soup.find_all('div', class_='item-img')
                    self.links_products[art] = found_links_some_product[0].find('a')['href']
                # когда ссылка на новую страницу с продуктом
                else:
                    article_on_page = soup.find('div', class_='goodsDataMainModificationArtNumber').find(
                        'span').text.strip()
                    # при совпадении искомого артикула с артикулом на странице
                    if art[3:] in article_on_page:
                        some_links_imgs = soup.find('div', class_='thumblist-box').find_all('li')
                        # при нескольких изображениях
                        if len(some_links_imgs) > 1:
                            self.article_imgs.setdefault(art, [img.find('a')['href'] for img in some_links_imgs])
                        elif len(some_links_imgs) == 1:
                            link_img_found = soup.find('div', class_='product-img-box col-md-5 col-sm-12 col-sms-12')
                            link_img = link_img_found.find_all('a')
                            self.article_imgs.setdefault(art, [link['href'] for link in link_img])

                        # при одном изображении
                        else:
                            link_img_found = soup.find('div', class_='general-img popup-gallery')
                            if link_img_found.find('a') == None:
                                continue
                            else:
                                link_img = link_img_found.find_all('a')
                                self.article_imgs.setdefault(art, [link['href'] for link in link_img])
                    # если на странице артикул не совпадает с искомым
                    else:
                        continue

            else:
                continue

    def get_link_img(self):

        for art in self.links_products:

            sys.stdout.write("\r")
            sys.stdout.write(f'Получение изображений для - {art}')
            sys.stdout.flush()

            resp = requests.get(self.links_products[art])
            soup = BeautifulSoup(resp.text, features='lxml')
            article_on_page = soup.find('div', class_='goodsDataMainModificationArtNumber').find(
                'span').text.strip()
            # если артикул на странице не совпадает с искомым
            if art[3:] not in article_on_page:
                continue
            else:
                some_links_imgs = soup.find('div', class_='thumblist-box').find_all('li')
                # при нескольких изображениях
                if len(some_links_imgs) > 1:
                    self.article_imgs.setdefault(art, [img.find('a')['href'] for img in some_links_imgs])
                elif len(some_links_imgs) == 1:
                    link_img_found = soup.find('div', class_='product-img-box col-md-5 col-sm-12 col-sms-12')
                    link_img = link_img_found.find_all('a')
                    self.article_imgs.setdefault(art, [link['href'] for link in link_img])

                # при одном изображении
                else:
                    link_img_found = soup.find('div', class_='general-img popup-gallery')
                    if link_img_found.find('a') == None:
                        continue
                    else:
                        link_img = link_img_found.find_all('a')
                        self.article_imgs.setdefault(art, [link['href'] for link in link_img])

    async def save_images(self, session, urls, name_img):
        self.article_imgs = {'BH-6624': [
            'https://i3.stat01.com/2/5417/154161236/afacdb/skovoroda-24-sm-berlinger-haus-bh-1857-royal-purple-metallic-line.jpg',
            'https://i4.stat01.com/2/5417/154161237/afacdb/skovoroda-24-sm-berlinger-haus-bh-1857-royal-purple-metallic-line.jpg'],
            'BH-6626': [
                'https://i3.stat01.com/2/5417/154161236/afacdb/skovoroda-24-sm-berlinger-haus-bh-1857-royal-purple-metallic-line.jpg',
                'https://i4.stat01.com/2/5417/154161237/afacdb/skovoroda-24-sm-berlinger-haus-bh-1857-royal-purple-metallic-line.jpg'],
            'BH-6629': [
                'https://i5.stat01.com/2/5094/150931582/afacdb/kastryulya-s-kryshkoj-20sm-berlinger-haus-vn-6628-purple-eclips-collection.jpg',
                'https://i4.stat01.com/2/5094/150931583/afacdb/kastryulya-s-kryshkoj-20sm-berlinger-haus-vn-6628-purple-eclips-collection.jpg'],
            'BH-6631': [
                'https://i1.stat01.com/2/4930/149297960/afacdb/sotejnik-s-kryshkoj-24sm-berlinger-haus-vn-6631-royal-purple-metallic-line.png',
                'https://i2.stat01.com/2/4930/149297961/afacdb/sotejnik-s-kryshkoj-24sm-berlinger-haus-vn-6631-royal-purple-metallic-line.png'],
            'BH-6633': [
                'https://i5.stat01.com/2/5417/154161207/afacdb/skovoroda-vok-28-sm-berlinger-haus-bh-1861-royal-purple-metallic-line.png',
                'https://i2.stat01.com/2/5417/154161208/afacdb/skovoroda-vok-28-sm-berlinger-haus-bh-1861-royal-purple-metallic-line.png'],
            'BH-6635': [
                'https://i2.stat01.com/2/7846/178451505/afacdb/blinnica-25sm-berlinger-haus-vn-1272-burgundy-metallic-line.jpg',
                'https://i4.stat01.com/2/7846/178451506/afacdb/blinnica-25sm-berlinger-haus-vn-1272-burgundy-metallic-line.jpg'],
            'BH-7131': [
                'https://i2.stat01.com/2/7846/178451505/afacdb/blinnica-25sm-berlinger-haus-vn-1272-burgundy-metallic-line.jpg',
                'https://i4.stat01.com/2/7846/178451506/afacdb/blinnica-25sm-berlinger-haus-vn-1272-burgundy-metallic-line.jpg'],
            'BH-7140': [
                'https://i2.stat01.com/2/8304/183038964/afacdb/bh-1660-purple-eclips-collection-nabor-posudy-10-pr-berlingerhaus.jpg',
                'https://i4.stat01.com/2/6526/165257115/afacdb/bh-1660-purple-eclips-collection-nabor-posudy-10-pr-berlingerhaus.jpg',
                'https://i4.stat01.com/2/6526/165257116/afacdb/bh-1660-purple-eclips-collection-nabor-posudy-10-pr-berlingerhaus.jpg',
                'https://i4.stat01.com/2/6526/165257117/afacdb/bh-1660-purple-eclips-collection-nabor-posudy-10-pr-berlingerhaus.jpg',
                'https://i4.stat01.com/2/6526/165257118/afacdb/bh-1660-purple-eclips-collection-nabor-posudy-10-pr-berlingerhaus.jpg',
                'https://i4.stat01.com/2/6526/165257119/afacdb/bh-1660-purple-eclips-collection-nabor-posudy-10-pr-berlingerhaus.jpg',
                'https://i4.stat01.com/2/8304/183038963/afacdb/bh-1660-purple-eclips-collection-nabor-posudy-10-pr-berlingerhaus.jpg'],
            'BH-7141': [
                'https://i2.stat01.com/2/8489/184888582/afacdb/bh-6152-aquamarine-edition-nabor-posudy-10-pr.jpg',
                'https://i1.stat01.com/2/8489/184888584/afacdb/bh-6152-aquamarine-edition-nabor-posudy-10-pr.jpg',
                'https://i2.stat01.com/2/8489/184888585/afacdb/bh-6152-aquamarine-edition-nabor-posudy-10-pr.jpg'],
            'BH-7143': [
                'https://i1.stat01.com/2/8489/184887874/afacdb/bh-7031-aquamarine-metallic-line-nabor-posudy-13-pr.jpg',
                'https://i1.stat01.com/2/8489/184887878/afacdb/bh-7031-aquamarine-metallic-line-nabor-posudy-13-pr.jpg',
                'https://i1.stat01.com/2/8489/184887879/afacdb/bh-7031-aquamarine-metallic-line-nabor-posudy-13-pr.jpg',
                'https://i1.stat01.com/2/8489/184887875/afacdb/bh-7031-aquamarine-metallic-line-nabor-posudy-13-pr.jpg',
                'https://i4.stat01.com/2/8489/184887876/afacdb/bh-7031-aquamarine-metallic-line-nabor-posudy-13-pr.jpg',
                'https://i2.stat01.com/2/8489/184887877/afacdb/bh-7031-aquamarine-metallic-line-nabor-posudy-13-pr.jpg',
                'https://i1.stat01.com/2/8489/184887871/afacdb/bh-7031-aquamarine-metallic-line-nabor-posudy-13-pr.jpg',
                'https://i2.stat01.com/2/8489/184887870/afacdb/bh-7031-aquamarine-metallic-line-nabor-posudy-13-pr.jpg',
                'https://i2.stat01.com/2/8489/184887873/afacdb/bh-7031-aquamarine-metallic-line-nabor-posudy-13-pr.jpg'],
            'BH-1870': [
                'https://i5.stat01.com/2/200/101997695/afacdb/skovoroda-20-sm-berlinger-haus-vn-1870-aquamarine-metallic-line.jpg',
                'https://i5.stat01.com/2/200/101997695/afacdb/skovoroda-20-sm-berlinger-haus-vn-1870-aquamarine-metallic-line.jpg'],
            'BH-1878': [
                'https://i4.stat01.com/2/392/103915694/afacdb/kastryulya-s-kryshkoj-20-sm-berlinger-haus-bh-1878.jpg'],
            'BH-1879': [
                'https://i1.stat01.com/2/392/103916267/afacdb/kastryulya-s-kryshkoj-20-sm-berlinger-haus-bh-1878.jpg'],
            'BH-1988': [
                'https://i3.stat01.com/2/5460/154597674/afacdb/tazhin-berlinger-haus-bh-1987-metallic-line.jpg',
                'https://i3.stat01.com/2/5460/154597674/afacdb/tazhin-berlinger-haus-bh-1987-metallic-line.jpg'],
            'BH-6101': [
                'https://i5.stat01.com/2/2492/124915899/afacdb/nabor-posudy-dlya-prigotovleniya-12-pr-aquamarine-edition-bh-6101-s-antiprigarnym-pokrytiem.jpg',
                'https://i5.stat01.com/2/2492/124915900/afacdb/nabor-posudy-dlya-prigotovleniya-12-pr-aquamarine-edition-bh-6101-s-antiprigarnym-pokrytiem.jpg'],
            'BH-7028': [
                'https://i2.stat01.com/2/8612/186112274/afacdb/gril-skovoroda-s-kryshkoj-28sm-berlingerhaus-bh-1613n-burgundy-metallic-line.jpg',
                'https://i3.stat01.com/2/8612/186112275/afacdb/gril-skovoroda-s-kryshkoj-28sm-berlingerhaus-bh-1613n-burgundy-metallic-line.jpg'],
            'BH-1252': [
                'https://i4.stat01.com/2/2022/120215441/afacdb/skovoroda-24sm-berlinger-haus-vn-1252-burgundy-metallic-line.jpg',
                'https://i4.stat01.com/2/2022/120215442/afacdb/skovoroda-24sm-berlinger-haus-vn-1252-burgundy-metallic-line.jpg'],
            'BH-1256': [
                'https://i4.stat01.com/1/7279/72784298/afacdb/kastryulya-2-5l-20sm-berlinger-haus-vn-1256-burgundy-metallic-line.jpg',
                'https://i3.stat01.com/1/7279/72784297/afacdb/kastryulya-2-5l-20sm-berlinger-haus-vn-1256-burgundy-metallic-line.jpg'],
            'BH-1259': [
                'https://i2.stat01.com/2/6610/166097079/afacdb/skovoroda-s-kryshkoj-24sm-berlinger-haus-vn-1259-burgundy-metallic-line.jpg',
                'https://i3.stat01.com/2/6610/166097080/afacdb/skovoroda-s-kryshkoj-24sm-berlinger-haus-vn-1259-burgundy-metallic-line.jpg',
                'https://i3.stat01.com/2/6610/166097081/afacdb/skovoroda-s-kryshkoj-24sm-berlinger-haus-vn-1259-burgundy-metallic-line.jpg',
                'https://i2.stat01.com/2/6610/166097082/afacdb/skovoroda-s-kryshkoj-24sm-berlinger-haus-vn-1259-burgundy-metallic-line.jpg'],
            'BH-1262': [
                'https://i5.stat01.com/2/2022/120215461/afacdb/sotejnik-s-kryshkoj-24sm-berlinger-haus-vn-1262-burgundy-metallic-line.jpg',
                'https://i4.stat01.com/2/2022/120215462/afacdb/sotejnik-s-kryshkoj-24sm-berlinger-haus-vn-1262-burgundy-metallic-line.jpg'],
            'BH-1265': [
                'https://i4.stat01.com/2/2022/120215468/afacdb/skovoroda-vok-s-kryshkoj-28sm-berlinger-haus-vn-1265-burgundy-metallic-line.jpg',
                'https://i2.stat01.com/2/2022/120215471/afacdb/skovoroda-vok-s-kryshkoj-28sm-berlinger-haus-vn-1265-burgundy-metallic-line.jpg'],
            'BH-1668': ['https://i4.stat01.com/1/9526/95250369/afacdb/vn-1668-jpg.jpg'], 'BH-1945': [
                'https://i3.stat01.com/2/4930/149296329/afacdb/skovoroda-24sm-berlinger-haus-vn-1252-burgundy-metallic-line.jpg',
                'https://i2.stat01.com/2/4930/149296330/afacdb/skovoroda-24sm-berlinger-haus-vn-1252-burgundy-metallic-line.jpg'],
            'BH-6150': [
                'https://i4.stat01.com/2/3589/135888777/afacdb/bh-6154-black-rose-nabor-posudy-10-pr.jpg',
                'https://i4.stat01.com/2/3589/135888778/afacdb/bh-6154-black-rose-nabor-posudy-10-pr.jpg'],
            'BH-6179': [
                'https://i4.stat01.com/2/2022/120215487/afacdb/blinnica-25sm-berlinger-haus-vn-1272-burgundy-metallic-line.jpg',
                'https://i4.stat01.com/2/2022/120215488/afacdb/blinnica-25sm-berlinger-haus-vn-1272-burgundy-metallic-line.jpg',
                'https://i5.stat01.com/2/2022/120215489/afacdb/blinnica-25sm-berlinger-haus-vn-1272-burgundy-metallic-line.jpg',
                'https://i2.stat01.com/2/2022/120215490/afacdb/blinnica-25sm-berlinger-haus-vn-1272-burgundy-metallic-line.jpg',
                'https://i2.stat01.com/2/2022/120215492/afacdb/blinnica-25sm-berlinger-haus-vn-1272-burgundy-metallic-line.jpg',
                'https://i5.stat01.com/2/3708/137077421/afacdb/blinnica-25sm-berlinger-haus-vn-1272-burgundy-metallic-line.jpg'],
            'BH-1990': ['https://i1.stat01.com/2/3731/137309925/afacdb/bh-1990-emerald-collection-tazhin.jpg',
                        'https://i1.stat01.com/2/3731/137309925/afacdb/bh-1990-emerald-collection-tazhin.jpg'],
            'BH-6047': [
                'https://i5.stat01.com/2/2561/125602697/afacdb/skovoroda-24sm-berlingerhaus-vn-1634n-black-rose.jpg',
                'https://i3.stat01.com/2/2561/125602730/afacdb/skovoroda-24sm-berlingerhaus-vn-1634n-black-rose.jpg'],
            'BH-6049': [
                'https://i5.stat01.com/2/2552/125514896/afacdb/skovoroda-s-kryshkoj-24-sm-berlinger-haus-bh-1690-bronze-titan.jpg',
                'https://i4.stat01.com/2/2552/125514918/afacdb/skovoroda-s-kryshkoj-24-sm-berlinger-haus-bh-1690-bronze-titan.jpg'],
            'BH-6055': [
                'https://i4.stat01.com/2/3444/134436954/afacdb/kovsh-16sm-berlingerhaus-vn-1624n-black-burgundy-metallic-line.jpg',
                'https://i4.stat01.com/2/3444/134436955/afacdb/kovsh-16sm-berlingerhaus-vn-1624n-black-burgundy-metallic-line.jpg',
                'https://i4.stat01.com/2/3444/134436956/afacdb/kovsh-16sm-berlingerhaus-vn-1624n-black-burgundy-metallic-line.jpg'],
            'BH-6057': [
                'https://i5.stat01.com/2/2561/125606262/afacdb/kastryulya-so-steklyannoj-kryshkoj-20sm-2-5l-berlingerhaus-vn-1641n-black-rose.jpg',
                'https://i2.stat01.com/2/2561/125606263/afacdb/kastryulya-so-steklyannoj-kryshkoj-20sm-2-5l-berlingerhaus-vn-1641n-black-rose.jpg',
                'https://i2.stat01.com/2/2561/125606264/afacdb/kastryulya-so-steklyannoj-kryshkoj-20sm-2-5l-berlingerhaus-vn-1641n-black-rose.jpg'],
            'BH-6035': [
                'https://i4.stat01.com/2/4516/145152503/afacdb/kastryulya-20sm-berlinger-haus-irose-6034-bh.jpg',
                'https://i4.stat01.com/2/4516/145152504/afacdb/kastryulya-20sm-berlinger-haus-irose-6034-bh.jpg'],
            'BH-6037': [
                'https://i3.stat01.com/2/2779/127782478/afacdb/sotejnik-s-kryshkoj-28sm-berlinger-haus-vn-6014-moonlight.jpg',
                'https://i5.stat01.com/2/2779/127782489/afacdb/sotejnik-s-kryshkoj-28sm-berlinger-haus-vn-6014-moonlight.jpg'],
            'BH-1989': ['https://i4.stat01.com/2/5988/159875534/afacdb/bh-1990-emerald-collection-tazhin.jpg',
                        'https://i4.stat01.com/2/5988/159875535/afacdb/bh-1990-emerald-collection-tazhin.jpg',
                        'https://i4.stat01.com/2/5988/159875537/afacdb/bh-1990-emerald-collection-tazhin.jpg'],
            'BH-6014': [
                'https://i2.stat01.com/2/5519/155188509/afacdb/sotejnik-s-kryshkoj-28sm-berlinger-haus-vn-6014-moonlight.jpg',
                'https://i3.stat01.com/2/5519/155188510/afacdb/sotejnik-s-kryshkoj-28sm-berlinger-haus-vn-6014-moonlight.jpg',
                'https://i2.stat01.com/2/5519/155188511/afacdb/sotejnik-s-kryshkoj-28sm-berlinger-haus-vn-6014-moonlight.jpg',
                'https://i2.stat01.com/2/5519/155188512/afacdb/sotejnik-s-kryshkoj-28sm-berlinger-haus-vn-6014-moonlight.png'],
            'BH-7033': [
                'https://i4.stat01.com/2/5794/157933872/afacdb/bh-7033-moonlight-edition-nabor-posudy-13-pr.jpg',
                'https://i3.stat01.com/2/5597/155963759/afacdb/bh-7033-moonlight-edition-nabor-posudy-13-pr.jpg',
                'https://i2.stat01.com/2/5597/155963760/afacdb/bh-7033-moonlight-edition-nabor-posudy-13-pr.jpg',
                'https://i5.stat01.com/2/5597/155963755/afacdb/bh-7033-moonlight-edition-nabor-posudy-13-pr.jpg',
                'https://i5.stat01.com/2/5597/155963756/afacdb/bh-7033-moonlight-edition-nabor-posudy-13-pr.jpg',
                'https://i5.stat01.com/2/5597/155963757/afacdb/bh-7033-moonlight-edition-nabor-posudy-13-pr.jpg',
                'https://i4.stat01.com/2/5794/157933871/afacdb/bh-7033-moonlight-edition-nabor-posudy-13-pr.jpg',
                'https://i4.stat01.com/2/5597/155963758/afacdb/bh-7033-moonlight-edition-nabor-posudy-13-pr.png'],
            'BH-1669': [
                'https://i5.stat01.com/2/5601/156009212/afacdb/nabor-posudy-3pr-berlingerhaus-vn-1669-rosegold-line.jpg',
                'https://i5.stat01.com/2/5601/156009213/afacdb/nabor-posudy-3pr-berlingerhaus-vn-1669-rosegold-line.jpg',
                'https://i5.stat01.com/2/5601/156009214/afacdb/nabor-posudy-3pr-berlingerhaus-vn-1669-rosegold-line.jpg',
                'https://i5.stat01.com/2/5601/156009215/afacdb/nabor-posudy-3pr-berlingerhaus-vn-1669-rosegold-line.jpg'],
            'BH-6194': [
                'https://i5.stat01.com/2/5597/155963622/afacdb/bh-6194-rosegold-line-kastryulya-s-kryshkoj-30sm.jpg',
                'https://i4.stat01.com/2/5601/156009613/afacdb/bh-6194-rosegold-line-kastryulya-s-kryshkoj-30sm.jpg',
                'https://i5.stat01.com/2/5601/156009622/afacdb/bh-6194-rosegold-line-kastryulya-s-kryshkoj-30sm.jpg'],
            'BH-6608': [
                'https://i2.stat01.com/2/5526/155256651/afacdb/bh-6608-shiny-black-edition-sotejnik-24sm.jpg',
                'https://i2.stat01.com/2/5526/155256653/afacdb/bh-6608-shiny-black-edition-sotejnik-24sm.jpg'],
            'BH-1599': ['https://i4.stat01.com/1/9558/95571379/afacdb/bh-1599-jpg.jpg'], 'BH-1713': [
                'https://i5.stat01.com/2/2022/120215221/afacdb/skovoroda-24sm-berlingerhaus-bh-1713-forest-line.jpg',
                'https://i4.stat01.com/2/5989/159882647/afacdb/skovoroda-24sm-berlingerhaus-bh-1713-ebony-rosewood-collection-forest-line.jpg'],
            'BH-1718': ['https://i4.stat01.com/1/9558/95575050/afacdb/bh-1718-jpg.jpg',
                        'https://i1.stat01.com/2/5989/159880310/afacdb/glubokaya-skovoroda-s-kryshkoj-24sm-berlingerhaus-bh-1718-ebony-rosewood-collection-forest-line.jpg'],
            'BH-1098': [
                'https://i4.stat01.com/2/2492/124915485/afacdb/kastryulya-s-kryshkoj-8-5l-32sm-berlinger-haus-granit-diamond-line-bh-1098.jpg',
                'https://i4.stat01.com/2/2492/124915486/afacdb/kastryulya-s-kryshkoj-8-5l-32sm-berlinger-haus-granit-diamond-line-bh-1098.jpg'],
            'BH-1791': [
                'https://i5.stat01.com/2/236/102359327/afacdb/skovoroda-20-sm-berlinger-haus-bh-1843-black-silver-line.jpg',
                'https://i5.stat01.com/2/236/102359329/afacdb/skovoroda-20-sm-berlinger-haus-bh-1843-black-silver-line.jpg'],
            'BH-1794': [
                'https://i5.stat01.com/2/2022/120215247/afacdb/skovoroda-gril-28-sm-berlinger-haus-bh-1794-granit-diamond-line.jpg',
                'https://i4.stat01.com/2/2022/120215248/afacdb/skovoroda-gril-28-sm-berlinger-haus-bh-1794-granit-diamond-line.jpg'],
            'BH-1795': [
                'https://i5.stat01.com/2/236/102359741/afacdb/kovsh-16-sm-berlinger-haus-bh-1860-royal-purple-metallic-line.jpg'],
            'BH-1931': [
                'https://i5.stat01.com/2/2492/124915258/afacdb/vn-1931-black-rose-collection-kastryulya-so-steklyannoj-kryshkoj-28sm-6-1l.jpg',
                'https://i2.stat01.com/2/2492/124915259/afacdb/vn-1931-black-rose-collection-kastryulya-so-steklyannoj-kryshkoj-28sm-6-1l.jpg',
                'https://i4.stat01.com/2/2492/124915260/afacdb/vn-1931-black-rose-collection-kastryulya-so-steklyannoj-kryshkoj-28sm-6-1l.jpg'],
            'BH-6558': [
                'https://i2.stat01.com/2/5471/154706685/afacdb/kovsh-16-sm-berlinger-haus-bh-1874-aquamarine-metallic-line.jpg',
                'https://i2.stat01.com/2/5471/154706686/afacdb/kovsh-16-sm-berlinger-haus-bh-1874-aquamarine-metallic-line.jpg',
                'https://i2.stat01.com/2/5471/154706687/afacdb/kovsh-16-sm-berlinger-haus-bh-1874-aquamarine-metallic-line.jpg'],
            'BH-6563': [
                'https://i1.stat01.com/2/5471/154706897/afacdb/bh-1877-aquamarine-metallic-line-sotejnik-s-kryshkoj-28sm.jpg',
                'https://i2.stat01.com/2/5471/154706898/afacdb/bh-1877-aquamarine-metallic-line-sotejnik-s-kryshkoj-28sm.jpg'],
            'BH-6582': [
                'https://i2.stat01.com/2/5471/154707007/afacdb/blinnica-25sm-berlinger-haus-vn-1876-aquamarine-metallic-line.jpg',
                'https://i5.stat01.com/2/5471/154707008/afacdb/blinnica-25sm-berlinger-haus-vn-1876-aquamarine-metallic-line.jpg',
                'https://i3.stat01.com/2/5471/154707009/afacdb/blinnica-25sm-berlinger-haus-vn-1876-aquamarine-metallic-line.jpg'],
            'BH-7430': [
                'https://i1.stat01.com/2/5471/154709257/afacdb/skovoroda-gril-28sm-berlinger-haus-vn-1271-burgundy-metallic-line.png'],
            'BH-6073': [
                'https://i4.stat01.com/2/8726/187253763/afacdb/bh-6073-primal-gloss-collection-nabor-posudy-3-pr.jpg',
                'https://i2.stat01.com/2/8726/187253764/afacdb/bh-6073-primal-gloss-collection-nabor-posudy-3-pr.jpg'],
            'BH-6568': [
                'https://i5.stat01.com/2/5244/152434000/afacdb/skovoroda-24sm-berlingerhaus-vn-6568-prima-gloss.png',
                'https://i5.stat01.com/2/5244/152434002/afacdb/skovoroda-24sm-berlingerhaus-vn-6568-prima-gloss.jpg'],
            'BH-6570': [
                'https://i3.stat01.com/2/5527/155261338/afacdb/bh-6570-primal-gloss-collection-kovsh-16sm.png'],
            'BH-6573': [
                'https://i3.stat01.com/2/5527/155261341/afacdb/bh-6573-primal-gloss-collection-kastryulya-so-steklyannoj-kryshkoj-28-sm.jpg'],
            'BH-6574': [
                'https://i3.stat01.com/2/5527/155261342/afacdb/bh-6574-primal-gloss-collection-skovoroda-so-steklyannoj-kryshkoj-24sm.png'],
            'BH-6578': [
                'https://i4.stat01.com/2/5526/155256737/afacdb/skovoroda-gril-berlinger-haus-bh-6578-primal-gloss-collection.jpg',
                'https://i4.stat01.com/2/5526/155256739/afacdb/skovoroda-gril-berlinger-haus-bh-6578-primal-gloss-collection.jpg'],
            'BH-9093': [
                'https://i4.stat01.com/2/6437/164366367/afacdb/bh-9093-jedektricheskij-chnik-purple-eclips.jpg',
                'https://i4.stat01.com/2/6439/164387659/afacdb/bh-9093-jedektricheskij-chnik-purple-eclips.jpg'],
            'BH-7217': [
                'https://i4.stat01.com/2/4615/146149501/afacdb/black-silver-line-jelektricheskaya-percemolka-1904-bh.jpg',
                'https://i4.stat01.com/2/4615/146149502/afacdb/black-silver-line-jelektricheskaya-percemolka-1904-bh.jpg'],
            'BH-2567': [
                'https://i3.stat01.com/2/6041/160403353/afacdb/bh-2567-carbon-metallic-line-nabor-nozhej-6pr.jpg',
                'https://i4.stat01.com/2/6053/160527399/afacdb/bh-2567-carbon-metallic-line-nabor-nozhej-6pr.jpg'],
            'BH-2461': [
                'https://i4.stat01.com/2/2806/128050548/afacdb/vn-2460-nabor-nozhej-na-podstavke-8-pr-aquamarine-metallic-line.jpg',
                'https://i1.stat01.com/2/2806/128050549/afacdb/vn-2460-nabor-nozhej-na-podstavke-8-pr-aquamarine-metallic-line.jpg'],
            'BH-2584': [
                'https://i4.stat01.com/2/6439/164387919/afacdb/bh-2584-purple-eclips-nabor-nozhej-na-podstavke-7-pr.jpg',
                'https://i4.stat01.com/2/6439/164387920/afacdb/bh-2584-purple-eclips-nabor-nozhej-na-podstavke-7-pr.jpg',
                'https://i4.stat01.com/2/6439/164387921/afacdb/bh-2584-purple-eclips-nabor-nozhej-na-podstavke-7-pr.jpg'],
            'BH-2600': [
                'https://i4.stat01.com/2/6657/166563037/afacdb/bh-2481-black-rose-collection-nabor-nozhej-na-podstavke-7-pr.jpg',
                'https://i5.stat01.com/2/6657/166563038/afacdb/bh-2481-black-rose-collection-nabor-nozhej-na-podstavke-7-pr.jpg'],
            'BH-2153': [
                'https://i5.stat01.com/2/2022/120215323/afacdb/stolovye-pribory-24pr-berlinger-haus-bh-2153-black-royal.jpg',
                'https://i5.stat01.com/2/2022/120215324/afacdb/stolovye-pribory-24pr-berlinger-haus-bh-2153-black-royal.jpg',
                'https://i5.stat01.com/2/2022/120215325/afacdb/stolovye-pribory-24pr-berlinger-haus-bh-2153-black-royal.png'],
            'BH-2156': [
                'https://i5.stat01.com/2/2022/120215330/afacdb/stolovye-pribory-24pr-berlinger-haus-bh-2156-black-royal.jpg',
                'https://i4.stat01.com/2/2022/120215331/afacdb/stolovye-pribory-24pr-berlinger-haus-bh-2156-black-royal.png'],
            'BH-2617': ['https://i5.stat01.com/2/4034/140334245/afacdb/bh-2617lifestyle02-jpg.jpg',
                        'https://i4.stat01.com/2/4034/140334246/afacdb/bh-2617-2-jpg.jpg',
                        'https://i4.stat01.com/2/4034/140334257/afacdb/untitled-design-2-jpg.jpg'],
            'BH-2621': ['https://i5.stat01.com/2/4034/140334576/afacdb/bh-2621lifestyle02-600x605-jpg.jpg',
                        'https://i5.stat01.com/2/4034/140334577/afacdb/bh-2621-1-jpg.jpg',
                        'https://i5.stat01.com/2/4034/140334578/afacdb/bh-2621-5-600x600-jpg.jpg',
                        'https://i4.stat01.com/2/4034/140334579/afacdb/untitled-design-6-jpg.jpg'],
            'BH-2623': [
                'https://i4.stat01.com/2/4074/140731396/afacdb/nabor-stolovyh-priborov-24pr-berlinger-haus-rosegold-line-bh-2637.png',
                'https://i3.stat01.com/2/4074/140731397/afacdb/nabor-stolovyh-priborov-24pr-berlinger-haus-rosegold-line-bh-2637.png'],
            'BH-7081': [
                'https://i3.stat01.com/2/8491/184902464/afacdb/nabor-posudy-10pr-berlingerhaus-vn-1645n-black-rose.jpg',
                'https://i3.stat01.com/2/8491/184902466/afacdb/nabor-posudy-10pr-berlingerhaus-vn-1645n-black-rose.jpg',
                'https://i3.stat01.com/2/8491/184902467/afacdb/nabor-posudy-10pr-berlingerhaus-vn-1645n-black-rose.jpg',
                'https://i4.stat01.com/2/8491/184902468/afacdb/nabor-posudy-10pr-berlingerhaus-vn-1645n-black-rose.png',
                'https://i3.stat01.com/2/8491/184902469/afacdb/nabor-posudy-10pr-berlingerhaus-vn-1645n-black-rose.jpg'],
            'BH-7082': [
                'https://i3.stat01.com/2/8490/184895455/afacdb/skovoroda-20-sm-berlinger-haus-bh-6624-royal-purple-metallic-line.jpg',
                'https://i3.stat01.com/2/8490/184895456/afacdb/skovoroda-20-sm-berlinger-haus-bh-6624-royal-purple-metallic-line.jpg'],
            'BH-7083': [
                'https://i3.stat01.com/2/8490/184896031/afacdb/skovoroda-24sm-berlingerhaus-vn-1634n-black-rose.jpg',
                'https://i4.stat01.com/2/8490/184896032/afacdb/skovoroda-24sm-berlingerhaus-vn-1634n-black-rose.jpg'],
            'BH-7088': [
                'https://i2.stat01.com/2/8491/184901285/afacdb/vn-1931-black-rose-collection-kastryulya-so-steklyannoj-kryshkoj-28sm-6-1l.jpg',
                'https://i1.stat01.com/2/8491/184901287/afacdb/vn-1931-black-rose-collection-kastryulya-so-steklyannoj-kryshkoj-28sm-6-1l.jpg'],
            'BH-6625': [
                'https://i3.stat01.com/2/5417/154161236/afacdb/skovoroda-24-sm-berlinger-haus-bh-1857-royal-purple-metallic-line.jpg',
                'https://i4.stat01.com/2/5417/154161237/afacdb/skovoroda-24-sm-berlinger-haus-bh-1857-royal-purple-metallic-line.jpg'],
            'BH-6627': [
                'https://i3.stat01.com/2/5540/155391320/afacdb/kovsh-16-sm-berlinger-haus-bh-1874-aquamarine-metallic-line.png',
                'https://i4.stat01.com/2/5540/155391321/afacdb/kovsh-16-sm-berlinger-haus-bh-1874-aquamarine-metallic-line.png'],
            'BH-6628': [
                'https://i5.stat01.com/2/5094/150931582/afacdb/kastryulya-s-kryshkoj-20sm-berlinger-haus-vn-6628-purple-eclips-collection.jpg',
                'https://i4.stat01.com/2/5094/150931583/afacdb/kastryulya-s-kryshkoj-20sm-berlinger-haus-vn-6628-purple-eclips-collection.jpg'],
            'BH-6630': [
                'https://i5.stat01.com/2/5094/150931582/afacdb/kastryulya-s-kryshkoj-20sm-berlinger-haus-vn-6628-purple-eclips-collection.jpg',
                'https://i4.stat01.com/2/5094/150931583/afacdb/kastryulya-s-kryshkoj-20sm-berlinger-haus-vn-6628-purple-eclips-collection.jpg'],
            'BH-6632': [
                'https://i1.stat01.com/2/4930/149297960/afacdb/sotejnik-s-kryshkoj-24sm-berlinger-haus-vn-6631-royal-purple-metallic-line.png',
                'https://i2.stat01.com/2/4930/149297961/afacdb/sotejnik-s-kryshkoj-24sm-berlinger-haus-vn-6631-royal-purple-metallic-line.png'],
            'BH-7026': [
                'https://i1.stat01.com/2/8490/184891863/afacdb/flip-skovoroda-26sm-berlinger-haus-vn-7026-purple-eclips-collection.jpg',
                'https://i4.stat01.com/2/8490/184891865/afacdb/flip-skovoroda-26sm-berlinger-haus-vn-7026-purple-eclips-collection.jpg'],
            'BH-7102': [
                'https://i1.stat01.com/2/8490/184893593/afacdb/nabor-posudy-5-pr-berlinger-haus-vn-6790-black-rose.jpg',
                'https://i3.stat01.com/2/8490/184893594/afacdb/nabor-posudy-5-pr-berlinger-haus-vn-6790-black-rose.jpg',
                'https://i4.stat01.com/2/8490/184893595/afacdb/nabor-posudy-5-pr-berlinger-haus-vn-6790-black-rose.jpg',
                'https://i3.stat01.com/2/8490/184893596/afacdb/nabor-posudy-5-pr-berlinger-haus-vn-6790-black-rose.jpg',
                'https://i3.stat01.com/2/8490/184893597/afacdb/nabor-posudy-5-pr-berlinger-haus-vn-6790-black-rose.jpg',
                'https://i1.stat01.com/2/8490/184893598/afacdb/nabor-posudy-5-pr-berlinger-haus-vn-6790-black-rose.jpg',
                'https://i3.stat01.com/2/8490/184893599/afacdb/nabor-posudy-5-pr-berlinger-haus-vn-6790-black-rose.jpg'],
            'BH-7103': [
                'https://i3.stat01.com/2/8489/184888885/afacdb/bh-6177-aquamarine-edition-nabor-posudy-3-pr.jpg',
                'https://i4.stat01.com/2/8489/184888883/afacdb/bh-6177-aquamarine-edition-nabor-posudy-3-pr.jpg',
                'https://i4.stat01.com/2/8489/184888886/afacdb/bh-6177-aquamarine-edition-nabor-posudy-3-pr.jpg',
                'https://i4.stat01.com/2/8489/184888887/afacdb/bh-6177-aquamarine-edition-nabor-posudy-3-pr.jpg',
                'https://i3.stat01.com/2/8489/184888888/afacdb/bh-6177-aquamarine-edition-nabor-posudy-3-pr.jpg'],
            'BH-7104': [
                'https://i4.stat01.com/2/8726/187253719/afacdb/bh-7104-purple-eclips-nabor-posudy-3-pr.jpg',
                'https://i1.stat01.com/2/8726/187253718/afacdb/bh-7104-purple-eclips-nabor-posudy-3-pr.jpg'],
            'BH-7105': [
                'https://i1.stat01.com/2/8490/184893617/afacdb/gril-skovoroda-s-kryshkoj-28sm-berlingerhaus-bh-1613n-burgundy-metallic-line.jpg',
                'https://i3.stat01.com/2/8514/185139235/afacdb/bh-7105-purple-eclips-gril-skovoroda-s-kryshkoj-28sm.jpg',
                'https://i2.stat01.com/2/8514/185139239/afacdb/bh-7105-purple-eclips-gril-skovoroda-s-kryshkoj-28sm.jpg'],
            'BH-7107': [
                'https://i2.stat01.com/2/8490/184893609/afacdb/bh-6189-burgundy-metallic-line-nabor-posudy-6-pr.jpg',
                'https://i1.stat01.com/2/8490/184893610/afacdb/bh-6189-burgundy-metallic-line-nabor-posudy-6-pr.jpg',
                'https://i1.stat01.com/2/8490/184893612/afacdb/bh-6189-burgundy-metallic-line-nabor-posudy-6-pr.jpg',
                'https://i1.stat01.com/2/8490/184893613/afacdb/bh-6189-burgundy-metallic-line-nabor-posudy-6-pr.jpg'],
            'BH-7138': [
                'https://i4.stat01.com/2/8489/184889583/afacdb/kovsh-16sm-berlingerhaus-bh-1525n-burgundy-metallic-line.jpg',
                'https://i1.stat01.com/2/8489/184889584/afacdb/kovsh-16sm-berlingerhaus-bh-1525n-burgundy-metallic-line.jpg',
                'https://i3.stat01.com/2/8489/184889585/afacdb/kovsh-16sm-berlingerhaus-bh-1525n-burgundy-metallic-line.jpg',
                'https://i3.stat01.com/2/8489/184889586/afacdb/kovsh-16sm-berlingerhaus-bh-1525n-burgundy-metallic-line.png'],
            'BH-7145': [
                'https://i1.stat01.com/2/8489/184889234/afacdb/bh-7145-purple-eclips-nabor-posudy-18pr.jpg',
                'https://i3.stat01.com/2/8489/184889235/afacdb/bh-7145-purple-eclips-nabor-posudy-18pr.jpg',
                'https://i2.stat01.com/2/8489/184883712/afacdb/bh-7036-aquamarine-metallic-line-nabor-posudy-18pr.jpg',
                'https://i1.stat01.com/2/8489/184883713/afacdb/bh-7036-aquamarine-metallic-line-nabor-posudy-18pr.png',
                'https://i4.stat01.com/2/8489/184883784/afacdb/bh-7036-aquamarine-metallic-line-nabor-posudy-18pr.jpg',
                'https://i3.stat01.com/2/8489/184883783/afacdb/bh-7036-aquamarine-metallic-line-nabor-posudy-18pr.jpg',
                'https://i4.stat01.com/2/8489/184883782/afacdb/bh-7036-aquamarine-metallic-line-nabor-posudy-18pr.jpg',
                'https://i4.stat01.com/2/8489/184883781/afacdb/bh-7036-aquamarine-metallic-line-nabor-posudy-18pr.jpg',
                'https://i1.stat01.com/2/8489/184883780/afacdb/bh-7036-aquamarine-metallic-line-nabor-posudy-18pr.png',
                'https://i1.stat01.com/2/8489/184883710/afacdb/bh-7036-aquamarine-metallic-line-nabor-posudy-18pr.jpg'],
            'BH-1884': [
                'https://i5.stat01.com/2/2022/120215635/afacdb/nabor-posudy-10-pr-berlingerhaus-vn-1884-aquamarine-metallic-line.jpg'],
            'BH-6143': [
                'https://i3.stat01.com/2/2852/128519417/afacdb/nabor-posudy-10-pr-berlingerhaus-vn-1884-aquamarine-metallic-line.jpg',
                'https://i5.stat01.com/2/2852/128519420/afacdb/nabor-posudy-10-pr-berlingerhaus-vn-1884-aquamarine-metallic-line.jpg',
                'https://i4.stat01.com/2/2852/128519480/afacdb/bh-6143-aquamarine-edition-nabor-posudy-10-pr-berlingerhaus.jpg',
                'https://i4.stat01.com/2/6157/161560350/afacdb/bh-6143-aquamarine-edition-nabor-posudy-10-pr-berlingerhaus.jpg',
                'https://i2.stat01.com/2/6157/161560351/afacdb/bh-6143-aquamarine-edition-nabor-posudy-10-pr-berlingerhaus.jpg'],
            'BH-6152': [
                'https://i4.stat01.com/2/3445/134446812/afacdb/bh-6151-rosegold-line-nabor-posudy-10-pr.jpg',
                'https://i4.stat01.com/2/3445/134446811/afacdb/bh-6151-rosegold-line-nabor-posudy-10-pr.jpg',
                'https://i5.stat01.com/2/3445/134446988/afacdb/bh-6151-rosegold-line-nabor-posudy-10-pr.png',
                'https://i4.stat01.com/2/6157/161560355/afacdb/bh-6152-aquamarine-edition-nabor-posudy-10-pr.jpg'],
            'BH-6165': [
                'https://i5.stat01.com/2/200/101997723/afacdb/skovoroda-s-kryshkoj-28-sm-berlinger-haus-bh-1877-aquamarine-metallic-line.jpg',
                'https://i5.stat01.com/2/200/101997723/afacdb/skovoroda-s-kryshkoj-28-sm-berlinger-haus-bh-1877-aquamarine-metallic-line.jpg'],
            'BH-1626N': [
                'https://i4.stat01.com/2/5894/158931737/afacdb/sotejnik-s-kryshkoj-28sm-berlingerhaus-vn-1626n-black-burgundy-metallic-line.jpg',
                'https://i4.stat01.com/2/5894/158931738/afacdb/sotejnik-s-kryshkoj-28sm-berlingerhaus-vn-1626n-black-burgundy-metallic-line.jpg'],
            'BH-1251': [
                'https://i4.stat01.com/2/2022/120215439/afacdb/skovoroda-20sm-berlinger-haus-vn-1251-burgundy-metallic-line.jpg',
                'https://i4.stat01.com/2/2022/120215440/afacdb/skovoroda-20sm-berlinger-haus-vn-1251-burgundy-metallic-line.jpg'],
            'BH-1253': [
                'https://i4.stat01.com/2/2022/120215443/afacdb/skovoroda-28sm-berlinger-haus-vn-1253-burgundy-metallic-line.jpg',
                'https://i4.stat01.com/2/2022/120215444/afacdb/skovoroda-28sm-berlinger-haus-vn-1253-burgundy-metallic-line.jpg'],
            'BH-1254': [
                'https://i4.stat01.com/1/7279/72789955/afacdb/skovoroda-30sm-berlinger-haus-vn-1254-carbon-metallic-line.jpg',
                'https://i4.stat01.com/1/7279/72789954/afacdb/skovoroda-30sm-berlinger-haus-vn-1254-carbon-metallic-line.jpg'],
            'BH-1255': [
                'https://i3.stat01.com/1/7279/72784541/afacdb/kovsh-1-3l-16sm-berlinger-haus-vn-1255-carbon-metallic-line.jpg',
                'https://i4.stat01.com/1/7279/72784540/afacdb/kovsh-1-3l-16sm-berlinger-haus-vn-1255-carbon-metallic-line.jpg'],
            'BH-1257': ['https://i5.stat01.com/2/2147/121465965/afacdb/screenshot6-png.png',
                        'https://i4.stat01.com/2/2147/121465966/afacdb/screenshot7-png.png'], 'BH-1258': [
                'https://i4.stat01.com/2/2492/124915393/afacdb/kastryulya-28sm-berlinger-haus-vn-1258-burgundy-metallic-line.jpg',
                'https://i5.stat01.com/2/2492/124915394/afacdb/kastryulya-28sm-berlinger-haus-vn-1258-burgundy-metallic-line.jpg'],
            'BH-1260': [
                'https://i4.stat01.com/2/2022/120215457/afacdb/skovoroda-s-kryshkoj-28sm-berlinger-haus-vn-1260-burgundy-metallic-line.jpg',
                'https://i4.stat01.com/2/2022/120215458/afacdb/skovoroda-s-kryshkoj-28sm-berlinger-haus-vn-1260-burgundy-metallic-line.jpg'],
            'BH-1261': [
                'https://i5.stat01.com/2/2022/120215459/afacdb/skovoroda-s-kryshkoj-32sm-berlinger-haus-vn-1261-burgundy-metallic-line.jpg',
                'https://i5.stat01.com/2/2022/120215460/afacdb/skovoroda-s-kryshkoj-32sm-berlinger-haus-vn-1261-burgundy-metallic-line.jpg'],
            'BH-1264': [
                'https://i4.stat01.com/1/7280/72791304/afacdb/sotejnik-s-kryshkoj-32sm-berlinger-haus-vn-1264-carbon-metallic-line.jpg',
                'https://i3.stat01.com/1/7280/72791303/afacdb/sotejnik-s-kryshkoj-32sm-berlinger-haus-vn-1264-carbon-metallic-line.jpg'],
            'BH-1267': [
                'https://i4.stat01.com/1/7279/72789963/afacdb/skovoroda-vok-28sm-berlinger-haus-vn-1267-burgundy-metallic-line.jpg',
                'https://i3.stat01.com/1/7279/72789962/afacdb/skovoroda-vok-28sm-berlinger-haus-vn-1267-burgundy-metallic-line.jpg'],
            'BH-1613N': [
                'https://i5.stat01.com/2/2061/120601387/afacdb/gril-skovoroda-s-kryshkoj-28sm-berlingerhaus-bh-1613n-burgundy-metallic-line.jpg',
                'https://i2.stat01.com/2/2061/120601388/afacdb/gril-skovoroda-s-kryshkoj-28sm-berlingerhaus-bh-1613n-burgundy-metallic-line.jpg'],
            'BH-1674': [
                'https://i2.stat01.com/2/2492/124915802/afacdb/nabor-posudy-12-pr-berlinger-haus-metallic-line-burgundy-edition-vn-1674.png',
                'https://i4.stat01.com/2/2492/124915803/afacdb/nabor-posudy-12-pr-berlinger-haus-metallic-line-burgundy-edition-vn-1674.png'],
            'BH-6166': [
                'https://i2.stat01.com/2/2492/124916093/afacdb/skovoroda-vok-28sm-berlinger-haus-s-burgundy-metallic-line.jpg',
                'https://i4.stat01.com/2/2492/124916094/afacdb/skovoroda-vok-28sm-berlinger-haus-s-burgundy-metallic-line.jpg'],
            'BH-7035': [
                'https://i4.stat01.com/2/5595/155947600/afacdb/bh-7035-burgundy-metallic-line-nabor-posudy-18pr.jpg',
                'https://i3.stat01.com/2/5595/155947602/afacdb/bh-7035-burgundy-metallic-line-nabor-posudy-18pr.jpg',
                'https://i5.stat01.com/2/5595/155947603/afacdb/bh-7035-burgundy-metallic-line-nabor-posudy-18pr.jpg',
                'https://i3.stat01.com/2/5595/155947605/afacdb/bh-7035-burgundy-metallic-line-nabor-posudy-18pr.jpg',
                'https://i3.stat01.com/2/5595/155947606/afacdb/bh-7035-burgundy-metallic-line-nabor-posudy-18pr.jpg',
                'https://i3.stat01.com/2/5595/155947700/afacdb/bh-7035-burgundy-metallic-line-nabor-posudy-18pr.jpg',
                'https://i5.stat01.com/2/5595/155947701/afacdb/bh-7035-burgundy-metallic-line-nabor-posudy-18pr.jpg',
                'https://i2.stat01.com/2/5595/155947702/afacdb/bh-7035-burgundy-metallic-line-nabor-posudy-18pr.jpg'],
            'BH-7110': [
                'https://i4.stat01.com/2/8672/186713899/afacdb/skovoroda-vok-s-kryshkoj-30sm-berlinger-haus-vn-1266-burgundy-metallic-line.jpg'],
            'BH-6918': [
                'https://i4.stat01.com/2/6041/160403376/afacdb/kontejner-d-syp-prod-wr-6918-1-41l.png'],
            'BH-6051': [
                'https://i3.stat01.com/2/4780/147796855/afacdb/bh-6051-emerald-collection-metallic-line-gril-skovoroda-so-steklyannoj-kryshkoj-28sm.jpg',
                'https://i3.stat01.com/2/4780/147796856/afacdb/bh-6051-emerald-collection-metallic-line-gril-skovoroda-so-steklyannoj-kryshkoj-28sm.jpg',
                'https://i3.stat01.com/2/4780/147796857/afacdb/bh-6051-emerald-collection-metallic-line-gril-skovoroda-so-steklyannoj-kryshkoj-28sm.jpg'],
            'BH-6087': [
                'https://i4.stat01.com/2/6439/164385295/afacdb/bh-6087-skovoroda-24sm-berlinger-haus-emerald-collection.jpg',
                'https://i4.stat01.com/2/6439/164385296/afacdb/bh-6087-skovoroda-24sm-berlinger-haus-emerald-collection.jpg'],
            'BH-6024': [
                'https://i5.stat01.com/2/2202/122010750/afacdb/skovoroda-berlinger-haus-irose-s-antiprigarnym-pokrytiem-diametr-28-sm-6025-bh.png',
                'https://i2.stat01.com/2/2202/122010751/afacdb/skovoroda-berlinger-haus-irose-s-antiprigarnym-pokrytiem-diametr-28-sm-6025-bh.png'],
            'BH-6104': ['https://i4.stat01.com/2/2231/122300396/afacdb/screenshot7-png.png',
                        'https://i3.stat01.com/2/2231/122300398/afacdb/screenshot6-png.png'], 'BH-6012': [
                'https://i5.stat01.com/2/2492/124915392/afacdb/kastryulya-28sm-berlinger-haus-moonlight-edition-6013-vn.png',
                'https://i5.stat01.com/2/5519/155188494/afacdb/kastryulya-24-sm-berlinger-haus-moonlight-edition-6012-vn.png',
                'https://i4.stat01.com/2/5519/155188495/afacdb/kastryulya-24-sm-berlinger-haus-moonlight-edition-6012-vn.jpg',
                'https://i2.stat01.com/2/5519/155188499/afacdb/kastryulya-24-sm-berlinger-haus-moonlight-edition-6012-vn.jpg',
                'https://i4.stat01.com/2/5519/155188500/afacdb/kastryulya-24-sm-berlinger-haus-moonlight-edition-6012-vn.png',
                'https://i4.stat01.com/2/5519/155188549/afacdb/kastryulya-24-sm-berlinger-haus-moonlight-edition-6012-vn.jpg'],
            'BH-6169': [
                'https://i4.stat01.com/2/5519/155188407/afacdb/vn-6169-moonlight-edition-nabor-posudy-3-pr.jpg',
                'https://i4.stat01.com/2/6543/165427579/afacdb/vn-6169-moonlight-edition-nabor-posudy-3-pr.jpg',
                'https://i2.stat01.com/2/6543/165427578/afacdb/vn-6169-moonlight-edition-nabor-posudy-3-pr.jpg',
                'https://i3.stat01.com/2/6543/165427576/afacdb/vn-6169-moonlight-edition-nabor-posudy-3-pr.jpg',
                'https://i1.stat01.com/2/6543/165427577/afacdb/vn-6169-moonlight-edition-nabor-posudy-3-pr.jpg',
                'https://i3.stat01.com/2/5519/155188429/afacdb/vn-6169-moonlight-edition-nabor-posudy-3-pr.png'],
            'BH-1518N': [
                'https://i4.stat01.com/2/5602/156011145/afacdb/skovoroda-s-kryshkoj-28-sm-berlinger-haus-bh-1518n-rosegold-line.jpg',
                'https://i1.stat01.com/2/5602/156011144/afacdb/skovoroda-s-kryshkoj-28-sm-berlinger-haus-bh-1518n-rosegold-line.jpg',
                'https://i5.stat01.com/2/5602/156011146/afacdb/skovoroda-s-kryshkoj-28-sm-berlinger-haus-bh-1518n-rosegold-line.jpg'],
            'BH-6142': [
                'https://i2.stat01.com/2/5601/156004778/afacdb/bh-6142-rosegold-line-nabor-posudy-10-pr-berlingerhaus.jpg',
                'https://i4.stat01.com/2/5601/156004780/afacdb/bh-6142-rosegold-line-nabor-posudy-10-pr-berlingerhaus.jpg',
                'https://i1.stat01.com/2/3887/138862631/afacdb/bh-6142-rosegold-line-nabor-posudy-10-pr-berlingerhaus.jpg',
                'https://i2.stat01.com/2/5601/156004781/afacdb/bh-6142-rosegold-line-nabor-posudy-10-pr-berlingerhaus.jpg'],
            'BH-1894': [
                'https://i2.stat01.com/2/395/103948216/afacdb/kastryulya-s-kryshkoj-20-sm-berlinger-haus-bh-1878.jpg'],
            'BH-1900': [
                'https://i4.stat01.com/2/2784/127834120/afacdb/kombajn-testomes-royalty-line-1900v-rl-pkm1900-7-0.jpg',
                'https://i1.stat01.com/2/2784/127834121/afacdb/kombajn-testomes-royalty-line-1900v-rl-pkm1900-7-0.jpg',
                'https://i3.stat01.com/2/2784/127834122/afacdb/kombajn-testomes-royalty-line-1900v-rl-pkm1900-7-0.jpg'],
            'BH-6605': [
                'https://i4.stat01.com/2/6440/164390026/afacdb/bh-6605-shiny-black-edition-kastryulya-so-steklyannoj-kryshkoj-24sm.jpg',
                'https://i4.stat01.com/2/6440/164390027/afacdb/bh-6605-shiny-black-edition-kastryulya-so-steklyannoj-kryshkoj-24sm.jpg',
                'https://i4.stat01.com/2/6440/164390028/afacdb/bh-6605-shiny-black-edition-kastryulya-so-steklyannoj-kryshkoj-24sm.jpg',
                'https://i4.stat01.com/2/6440/164390029/afacdb/bh-6605-shiny-black-edition-kastryulya-so-steklyannoj-kryshkoj-24sm.jpg'],
            'BH-6616': [
                'https://i2.stat01.com/2/5527/155261475/afacdb/bh-6616-shiny-black-edition-nabor-posudy-5pr.jpg',
                'https://i4.stat01.com/2/5527/155261476/afacdb/bh-6616-shiny-black-edition-nabor-posudy-5pr.jpg'],
            'BH-1202': [
                'https://i3.stat01.com/2/2492/124916135/afacdb/sotejnik-s-kryshkoj-32sm-berlinger-haus-bh-1202-forest-line.jpg',
                'https://i3.stat01.com/2/2492/124916136/afacdb/sotejnik-s-kryshkoj-32sm-berlinger-haus-bh-1202-forest-line.jpg'],
            'BH-1204': [
                'https://i4.stat01.com/2/2022/120214993/afacdb/skovoroda-vok-28sm-berlinger-haus-bh-1204-forest-line.jpg',
                'https://i4.stat01.com/2/2022/120214994/afacdb/skovoroda-vok-28sm-berlinger-haus-bh-1204-forest-line.jpg'],
            'BH-1210': [
                'https://i5.stat01.com/2/2747/127461021/afacdb/sotejnik-s-kryshkoj-24sm-berlinger-haus-bh-1210-forest-line.jpg'],
            'BH-1095': [
                'https://i2.stat01.com/1/7110/71090720/afacdb/kastryulya-s-kryshkoj-4-2l-24sm-berlinger-haus-bh-1095-granit-diamond-line.jpg',
                'https://i2.stat01.com/1/7110/71090721/afacdb/kastryulya-s-kryshkoj-4-2l-24sm-berlinger-haus-bh-1095-granit-diamond-line.jpg'],
            'BH-1096': [
                'https://i5.stat01.com/2/2492/124915467/afacdb/kastryulya-s-kryshkoj-5-8l-28sm-berlinger-haus-granit-diamond-line-bh-1096.jpg',
                'https://i5.stat01.com/2/2492/124915468/afacdb/kastryulya-s-kryshkoj-5-8l-28sm-berlinger-haus-granit-diamond-line-bh-1096.jpg'],
            'BH-1107': [
                'https://i2.stat01.com/1/7110/71091063/afacdb/skovoroda-so-sjemnoj-ruchkoj-24sm-berlinger-haus-bh-1107-granit-diamond-line.jpg',
                'https://i1.stat01.com/1/7110/71091062/afacdb/skovoroda-so-sjemnoj-ruchkoj-24sm-berlinger-haus-bh-1107-granit-diamond-line.jpg'],
            'BH-1446': ['https://i1.stat01.com/2/8726/187253331/afacdb/bh-1989-moonlight-edition-tazhin.jpg',
                        'https://i1.stat01.com/2/8726/187253332/afacdb/bh-1989-moonlight-edition-tazhin.jpg',
                        'https://i2.stat01.com/2/8726/187253333/afacdb/bh-1989-moonlight-edition-tazhin.jpg',
                        'https://i1.stat01.com/2/8726/187253334/afacdb/bh-1989-moonlight-edition-tazhin.jpg'],
            'BH-1634N': [
                'https://i2.stat01.com/2/5543/155427355/afacdb/skovoroda-24sm-berlingerhaus-vn-1634n-black-rose.jpg',
                'https://i4.stat01.com/2/2783/127820630/afacdb/skovoroda-24sm-berlingerhaus-vn-1634n-black-rose.jpg'],
            'BH-1645N': [
                'https://i4.stat01.com/2/2022/120215580/afacdb/nabor-posudy-10pr-berlingerhaus-vn-1645n-black-rose.jpg',
                'https://i4.stat01.com/2/5936/159350276/afacdb/nabor-posudy-10pr-berlingerhaus-vn-1645n-black-rose.jpg'],
            'BH-6149': [
                'https://i4.stat01.com/2/3587/135861815/afacdb/bh-6149-black-rose-nabor-posudy-10-pr-berlingerhaus.jpg',
                'https://i4.stat01.com/2/3587/135861816/afacdb/bh-6149-black-rose-nabor-posudy-10-pr-berlingerhaus.jpg'],
            'BH-6154': [
                'https://i4.stat01.com/2/3587/135861788/afacdb/bh-6154-black-rose-nabor-posudy-10-pr.jpg',
                'https://i4.stat01.com/2/3587/135861789/afacdb/bh-6154-black-rose-nabor-posudy-10-pr.jpg'],
            'BH-6766': [
                'https://i5.stat01.com/2/5094/150936407/afacdb/sotejnik-s-kryshkoj-28sm-berlinger-haus-vn-1263-burgundy-metallic-line.jpg',
                'https://i5.stat01.com/2/5094/150936407/afacdb/sotejnik-s-kryshkoj-28sm-berlinger-haus-vn-1263-burgundy-metallic-line.jpg'],
            'BH-7423': [
                'https://i1.stat01.com/2/5471/154708304/afacdb/bh-7422-eternal-collection-sotejnik-so-steklyannoj-kryshkoj-24sm.jpg'],
            'BH-6074': [
                'https://i3.stat01.com/2/8726/187254224/afacdb/bh-7102-purple-eclips-nabor-posudy-4-pr.jpg',
                'https://i1.stat01.com/2/8726/187253970/afacdb/bh-7102-purple-eclips-nabor-posudy-4-pr.jpg',
                'https://i2.stat01.com/2/8726/187253971/afacdb/bh-7102-purple-eclips-nabor-posudy-4-pr.jpg',
                'https://i1.stat01.com/2/8726/187253972/afacdb/bh-7102-purple-eclips-nabor-posudy-4-pr.jpg'],
            'BH-6567': [
                'https://i3.stat01.com/2/5527/155261337/afacdb/bh-6567-primal-gloss-collection-skovoroda-20sm.png'],
            'BH-6569': [
                'https://i2.stat01.com/2/4930/149297180/afacdb/skovoroda-28sm-berlinger-haus-vn-6569-prima-gloss.jpg',
                'https://i1.stat01.com/2/4930/149297181/afacdb/skovoroda-28sm-berlinger-haus-vn-6569-prima-gloss.jpg'],
            'BH-6571': [
                'https://i3.stat01.com/2/5527/155261339/afacdb/bh-6571-primal-gloss-collection-kastryulya-so-steklyannoj-20sm.jpg'],
            'BH-6572': [
                'https://i3.stat01.com/2/5527/155261340/afacdb/bh-6572-primal-gloss-collection-kastryulya-so-steklyannoj-kryshkoj-24-sm.jpg'],
            'BH-6576': [
                'https://i2.stat01.com/2/5527/155261379/afacdb/bh-6576-primal-gloss-collection-sotejnik-24sm.jpg',
                'https://i2.stat01.com/2/5527/155261376/afacdb/bh-6576-primal-gloss-collection-sotejnik-24sm.jpg',
                'https://i2.stat01.com/2/5527/155261375/afacdb/bh-6576-primal-gloss-collection-sotejnik-24sm.jpg',
                'https://i2.stat01.com/2/5527/155261378/afacdb/bh-6576-primal-gloss-collection-sotejnik-24sm.png'],
            'BH-6577': [
                'https://i2.stat01.com/2/4930/149297485/afacdb/skovoroda-28sm-berlinger-haus-vn-6569-prima-gloss.jpg',
                'https://i1.stat01.com/2/4930/149297486/afacdb/skovoroda-28sm-berlinger-haus-vn-6569-prima-gloss.jpg'],
            'BH-6579': [
                'https://i4.stat01.com/2/6531/165304306/afacdb/skovoroda-blinnaya-25-sm-berlinger-haus-bh-1523n-rosegold-line.jpg',
                'https://i4.stat01.com/2/6531/165304305/afacdb/skovoroda-blinnaya-25-sm-berlinger-haus-bh-1523n-rosegold-line.jpg'],
            'BH-7144': [
                'https://i2.stat01.com/2/8726/187253839/afacdb/bh-7031-aquamarine-metallic-line-nabor-posudy-13-pr.jpg',
                'https://i3.stat01.com/2/8726/187253840/afacdb/bh-7031-aquamarine-metallic-line-nabor-posudy-13-pr.jpg',
                'https://i3.stat01.com/2/8726/187253854/afacdb/bh-7031-aquamarine-metallic-line-nabor-posudy-13-pr.jpg'],
            'BH-9289': [
                'https://i2.stat01.com/2/5393/153929780/afacdb/bh-7217-burgundy-metallic-line-jelektricheskaya-percemolka-1-1.jpg',
                'https://i2.stat01.com/2/5393/153929781/afacdb/bh-7217-burgundy-metallic-line-jelektricheskaya-percemolka-1-1.jpg']}
        try:
            images = []

            sys.stdout.write("\r")
            sys.stdout.write(f'Сохраняю изображение для {name_img}')
            sys.stdout.flush()

            for a, url in enumerate(urls):
                date_now = datetime.datetime.now()
                async with aiofiles.open(f'./img/{name_img}_{date_now.strftime("%M%S%f")}_{a}.jpg', mode='wb') as f:
                    async with session.get(url) as response:
                        images.append(f'./img/{name_img}_{date_now.strftime("%M%S%f")}_{a}.jpg')
                        async for x in response.content.iter_chunked(1024):
                            await f.write(x)

            self.article_imgs[name_img] = images
        except Exception as exc:
            print(f'Ошибка {exc} в сохранении изображений товаров')
            with open('error.txt', 'a', encoding='utf-8') as file:

                file.write(f'{datetime.datetime.now().strftime("%d-%m-%y %H:%M")} '
                           f'Ошибка {exc} в сохранении изображений товаров, функция - save_images()\n')

    async def save_images_run_async(self):
        if not os.path.isdir('./img/'):
            os.mkdir('./img/')
        async with aiohttp.ClientSession() as session:
            tasks = []
            for link in self.article_imgs:
                # urls=self.article_imgs[link][:3] берёт только 3 изображения
                task = asyncio.create_task(self.save_images(session, urls=self.article_imgs[link][:3], name_img=link))
                tasks.append(task)
                await asyncio.gather(*tasks)

    def resize_img(self):
        try:
            ImageFile.LOAD_TRUNCATED_IMAGES = True
            fixed_height = 426
            for img_file in tqdm(os.listdir('./img/')):
                if img_file[-4:] == '.jpg':
                    img = Image.open(f'./img/{img_file}')
                    height_percent = (fixed_height / float(img.size[1]))
                    width_size = int((float(img.size[0]) * float(height_percent)))
                    new_image = img.resize((width_size, fixed_height))
                    new_image.save(f'./img/{img_file}')
        except Exception as exc:
            print(f'Ошибка {exc} в изменении разрешения изображений')
            with open('error.txt', 'a', encoding='utf-8') as file:
                file.write(f'{datetime.datetime.now().strftime("%d-%m-%y %H:%M")} '
                           f'Ошибка {exc} в изменении разрешения изображений, функция - resize_img()\n')

    def sending_to_fotohosting(self):
        self.active_token = self.token[0]
        self.active_secret_key = self.secret_key[0]
        headers = {
            'Authorization': f'TOKEN {self.active_token}',
        }
        for img_url in self.article_imgs:

            img_short_link = []

            sys.stdout.write("\r")
            sys.stdout.write(f'Загружаю изображение для - {img_url}')
            sys.stdout.flush()

            img_links = self.article_imgs[img_url]

            for img in img_links:

                try:
                    files = {
                        'image': open(img, 'rb'),
                        'secret_key': (None, self.active_secret_key),
                    }
                    response = requests.post('https://api.imageban.ru/v1', headers=headers, files=files)
                    if response.json()['status'] == 200:
                        img_short_link.append(f"[URL=https://imageban.ru][IMG]{response.json()['data']['link']}"
                                              f"[/IMG][/URL]")
                    else:
                        print(f'Не удалось загрузить {img}')
                        continue
                except KeyError:
                    print(f'{img_url} ошибка загрузки изображения - {response.json()["error"]["message"]}\n')
                    with open('error.txt', 'a', encoding='utf-8') as file:
                        file.write(f'{datetime.datetime.now().strftime("%d-%m-%y %H:%M")} '
                                   f'{img} ошибка загрузки изображения, функция - sending_to_fotohosting()\n')
                    if response.json()["error"]["message"] == 'File reception error':
                        continue
                    elif response.json()["error"]["message"] == \
                            'Exceeded the daily limit of uploaded images for your account':
                        print('Переключение на второй аккаунт')

                        self.active_token = self.token[1]
                        self.active_secret_key = self.secret_key[1]

                        files = {
                            'image': open(img, 'rb'),
                            'secret_key': (None, self.active_secret_key),
                        }
                        response = requests.post('https://api.imageban.ru/v1', headers=headers, files=files)
                        if response.json()['status'] == 200:
                            img_short_link.append(f"[URL=https://imageban.ru][IMG]{response.json()['data']['link']}"
                                                  f"[/IMG][/URL]")
                        else:
                            print(f'Не удалось загрузить {img}')
                    continue
                except FileNotFoundError:
                    continue
                self.article_save_imgs[img_url] = img_short_link

    def write_final_file(self):
        try:
            columns = ['N', 'O', 'P']
            wb = load_workbook(filename=self.read_data_file)
            ws = wb.active

            ws['N8'] = 'Ссылки на фотографии'
            date_now = datetime.datetime.now()
            for article in self.article_save_imgs:
                for i, link in enumerate(self.article_save_imgs[article]):
                    for row in ws.iter_cols(min_col=2, max_col=2, min_row=9):
                        for cell in row:
                            if cell.value.strip().split(' ')[0][3:] in article:
                                ws[f'{columns[i]}{cell.row}'] = link

            file_name = f'data_final_{date_now.strftime("%d-%m-%y_%H-%M")}.xlsx'
            wb.save(filename=file_name)
            shutil.rmtree('./img/')
            print(f'Файл {file_name} сохранён')
        except Exception as exc:
            print(f'Ошибка {exc} в записи итогового файла')
            with open('error.txt', 'a', encoding='utf-8') as file:
                file.write(f'{datetime.datetime.now().strftime("%d-%m-%y %H:%M")} '
                           f'Ошибка {exc} в записи итогового файла, функция - write_final_file()\n')

    def run(self):
        try:
            print('Начало работы')
            self.open_token_file()
            self.read_file()
            print('Получаю артикул товаров')
            self.get_article_number()
            print('\rАртикулы получил')
            print('---------------------------\n')
            print('Получаю ссылки на товары')
            self.get_link_prodicts()
            print('\nСсылки получены')
            print('---------------------------\n')
            print('Ищу изображения товаров')
            self.get_link_img()
            print('\nИзображения получены')
            print('---------------------------\n')
            print('Скачиваю изображения')
            # asyncio.run(self.save_images_run_async())
            # print('\nСкачивание завершено')
            # print('---------------------------\n')
            # print('Измененяю размер изображений')
            # self.resize_img()
            # print('\rРазмеры изменены')
            # print('---------------------------\n')
            # print('Загружаю изображения на фотохостинг')
            # self.sending_to_fotohosting()
            # print('\nЗагрузка завершена')
            # print('---------------------------\n')
            # print('Записываю в итоговый файл')
            # self.write_final_file()
            print('Работа завершена')
            print('Для выхода нажмите Enter')
            input()
            print('---------------------------\n')
        except Exception as exc:
            print(f'Произошла ошибка {exc}')
            print('Для выхода нажмите Enter')
            input()
            print('---------------------------\n')


def main():
    p = Parser()
    p.run()


if __name__ == '__main__':
    main()
