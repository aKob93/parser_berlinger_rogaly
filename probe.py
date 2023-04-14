# -*- coding: utf8 -*-
import os
import re
import time
import lxml
import shutil
import sys
import aiohttp
import asyncio
import aiofiles
import requests
from bs4 import BeautifulSoup
from fake_useragent import UserAgent
from openpyxl import load_workbook
from tqdm import tqdm
import datetime
from PIL import Image, ImageFile


class Parser:

    def __init__(self):
        ua = UserAgent()
        self.headers = {'user_agent': ua.random}
        self.token = ''
        self.secret_key = ''
        self.active_token = ''
        self.active_secret_key = ''
        self.base_url = 'https://berlinger-haus-shop.ru'
        self.article_numbers = []
        self.links_products = {}
        self.article_imgs = {}
        self.article_save_imgs = {}
        self.read_data1_file = ''
        self.read_data2_file = ''

    def open_token_file(self):
        try:
            with open('token.txt', 'r') as file:
                for i, line in enumerate(file):
                    if i == 0:
                        self.token = line.split('=')[1].strip().split(', ')
                    elif i == 1:
                        self.secret_key = line.split('=')[1].strip().split(', ')
        except Exception:
            print('Не удалось прочитать token или secret_key')
            raise IndexError

    def read_file(self):
        try:
            for file in os.listdir():
                if file[:6] == 'data1.':
                    print(f'Получаю артикул товаров из файла {file}')
                    self.read_data1_file = file
                    self.get_article_number_data1()
                elif file[:6] == 'data2.':
                    print(f'Получаю артикул товаров из файла {file}')
                    self.read_data2_file = file
                    self.get_article_number_data2()
        except Exception:
            print('Нет файла с именем data.')
            raise IndexError

    def get_article_number_data1(self):
        try:
            wb = load_workbook(filename=self.read_data1_file)
            sheets = wb.sheetnames
            ws = wb[sheets[0]]

            for row in ws.iter_cols(min_col=2, max_col=2, min_row=9):
                for cell in row:
                    if cell.value is None:
                        continue
                    # есть ли числа в строке
                    if re.search('\d+', cell.value.strip().split(' ')[0]):
                        self.article_numbers.append(cell.value.strip().split(' ')[0])

            self.article_numbers = list(dict.fromkeys(self.article_numbers))
        except Exception as exc:
            print(f'Ошибка {exc} в чтении табличного документа data1.xlsx')
            with open('error.txt', 'a', encoding='utf-8') as file:

                file.write(f'{datetime.datetime.now().strftime("%d-%m-%y %H:%M")} '
                           f'Ошибка {exc} в чтении табличного документа data1.xlsm, функция - get_article_number()\n')
            raise IndexError

    def get_article_number_data2(self):
        try:
            wb = load_workbook(filename=self.read_data2_file)
            sheets = wb.sheetnames
            ws = wb[sheets[0]]

            for row in ws.iter_cols(min_col=8, max_col=8, min_row=10):
                for cell in row:
                    if cell.value is None:
                        continue
                    # есть ли числа в строке
                    # if re.search('\d+', cell.value.strip().split(' ')[0]):
                    self.article_numbers.append(cell.value.strip())

        except Exception as exc:
            print(f'Ошибка {exc} в чтении табличного документа data2.xlsx')
            with open('error.txt', 'a', encoding='utf-8') as file:

                file.write(f'{datetime.datetime.now().strftime("%d-%m-%y %H:%M")} '
                           f'Ошибка {exc} в чтении табличного документа data2.xlsm, функция - get_article_number()\n')
            raise IndexError

    def get_link_prodicts(self):
        print(self.article_numbers)

        for art in self.article_numbers:

            sys.stdout.write("\r")
            sys.stdout.write(f'Обрабатываю - {art}')
            sys.stdout.flush()

            if len(art) == 1:
                continue
            response = requests.get(f'{self.base_url}/search?q={art[3:]}', headers=self.headers)
            soup = BeautifulSoup(response.text, features='lxml')

            product_not_found = soup.find('p', class_='warning')
            if bool(product_not_found) is False:

                if f'{self.base_url}/search?q={art[3:]}' == response.url:
                    found_links_some_product = soup.find_all('div', class_='item-img')
                    self.links_products[art] = found_links_some_product[0].find('a')['href']
                # когда ссылка на новую страницу с продуктом
                else:
                    try:
                        article_on_page = soup.find('div', class_='goodsDataMainModificationArtNumber').find(
                            'span').text.strip()
                    except Exception as exc:
                        print(f'ошибка {exc} art = {art}')
                        continue
                    # при совпадении искомого артикула с артикулом на странице
                    if art[3:] in article_on_page:
                        some_links_imgs = soup.find('div', class_='thumblist-box').find_all('li')
                        # при нескольких изображениях
                        if len(some_links_imgs) > 1:
                            self.article_imgs.setdefault(art, [img.find('a')['href'] for img in some_links_imgs])
                        elif len(some_links_imgs) == 1:
                            link_img_found = soup.find('div', class_='product-img-box col-md-5 col-sm-12 col-sms-12')
                            link_img = link_img_found.find_all('a')
                            self.article_imgs.setdefault(art, [link['href'] for link in link_img])

                        # при одном изображении
                        else:
                            link_img_found = soup.find('div', class_='general-img popup-gallery')
                            if link_img_found.find('a') == None:
                                continue
                            else:
                                link_img = link_img_found.find_all('a')
                                self.article_imgs.setdefault(art, [link['href'] for link in link_img])
                    # если на странице артикул не совпадает с искомым
                    else:
                        continue

            else:
                continue
        print(self.article_imgs)

    def get_link_img(self):

        for art in self.links_products:

            sys.stdout.write("\r")
            sys.stdout.write(f'Получение изображений для - {art}')
            sys.stdout.flush()

            resp = requests.get(self.links_products[art])
            soup = BeautifulSoup(resp.text, features='lxml')
            article_on_page = soup.find('div', class_='goodsDataMainModificationArtNumber').find(
                'span').text.strip()
            # если артикул на странице не совпадает с искомым
            if art[3:] not in article_on_page:
                continue
            else:
                some_links_imgs = soup.find('div', class_='thumblist-box').find_all('li')
                # при нескольких изображениях
                if len(some_links_imgs) > 1:
                    self.article_imgs.setdefault(art, [img.find('a')['href'] for img in some_links_imgs])
                elif len(some_links_imgs) == 1:
                    link_img_found = soup.find('div', class_='product-img-box col-md-5 col-sm-12 col-sms-12')
                    link_img = link_img_found.find_all('a')
                    self.article_imgs.setdefault(art, [link['href'] for link in link_img])

                # при одном изображении
                else:
                    link_img_found = soup.find('div', class_='general-img popup-gallery')
                    if link_img_found.find('a') == None:
                        continue
                    else:
                        link_img = link_img_found.find_all('a')
                        self.article_imgs.setdefault(art, [link['href'] for link in link_img])
        print(self.article_imgs)


    async def save_images(self, session, urls, name_img):

        try:
            images = []

            sys.stdout.write("\r")
            sys.stdout.write(f'Сохраняю изображение для {name_img}')
            sys.stdout.flush()

            for a, url in enumerate(urls):
                date_now = datetime.datetime.now()
                async with aiofiles.open(f'./img/{name_img}_{date_now.strftime("%M%S%f")}_{a}.jpg', mode='wb') as f:
                    async with session.get(url) as response:
                        images.append(f'./img/{name_img}_{date_now.strftime("%M%S%f")}_{a}.jpg')
                        async for x in response.content.iter_chunked(1024):
                            await f.write(x)

            self.article_imgs[name_img] = images
        except Exception as exc:
            print(f'Ошибка {exc} в сохранении изображений товаров')
            with open('error.txt', 'a', encoding='utf-8') as file:

                file.write(f'{datetime.datetime.now().strftime("%d-%m-%y %H:%M")} '
                           f'Ошибка {exc} в сохранении изображений товаров, функция - save_images()\n')

    async def save_images_run_async(self):
        if not os.path.isdir('./img/'):
            os.mkdir('./img/')
        async with aiohttp.ClientSession() as session:
            tasks = []
            for link in self.article_imgs:
                # urls=self.article_imgs[link][:3] берёт только 3 изображения
                task = asyncio.create_task(self.save_images(session, urls=self.article_imgs[link][:3], name_img=link))
                tasks.append(task)
                await asyncio.gather(*tasks)

    def resize_img(self):
        try:
            ImageFile.LOAD_TRUNCATED_IMAGES = True
            fixed_height = 426
            for img_file in tqdm(os.listdir('./img/')):
                if img_file[-4:] == '.jpg':
                    img = Image.open(f'./img/{img_file}')
                    if img.mode in ("RGBA", "P"):
                        img = img.convert("RGB")
                    height_percent = (fixed_height / float(img.size[1]))
                    width_size = int((float(img.size[0]) * float(height_percent)))
                    new_image = img.resize((width_size, fixed_height))
                    new_image.save(f'./img/{img_file}')
        except Exception as exc:
            print(f'Ошибка {exc} в изменении разрешения изображений')
            with open('error.txt', 'a', encoding='utf-8') as file:
                file.write(f'{datetime.datetime.now().strftime("%d-%m-%y %H:%M")} '
                           f'Ошибка {exc} в изменении разрешения изображений, функция - resize_img()\n')

    def sending_to_fotohosting(self):
        self.active_token = self.token[0]
        self.active_secret_key = self.secret_key[0]
        headers = {
            'Authorization': f'TOKEN {self.active_token}',
        }
        for img_url in self.article_imgs:

            img_short_link = []

            sys.stdout.write("\r")
            sys.stdout.write(f'Загружаю изображение для - {img_url}')
            sys.stdout.flush()

            img_links = self.article_imgs[img_url]

            for img in img_links:

                try:
                    files = {
                        'image': open(img, 'rb'),
                        'secret_key': (None, self.active_secret_key),
                    }
                    response = requests.post('https://api.imageban.ru/v1', headers=headers, files=files)
                    if response.json()['status'] == 200:
                        img_short_link.append(f"[URL=https://imageban.ru][IMG]{response.json()['data']['link']}"
                                              f"[/IMG][/URL]")
                    else:
                        print(f'Не удалось загрузить {img}')
                        continue
                except KeyError:
                    print(f'{img_url} ошибка загрузки изображения - {response.json()["error"]["message"]}\n')
                    with open('error.txt', 'a', encoding='utf-8') as file:
                        file.write(f'{datetime.datetime.now().strftime("%d-%m-%y %H:%M")} '
                                   f'{img} ошибка загрузки изображения, функция - sending_to_fotohosting()\n')
                    if response.json()["error"]["message"] == 'File reception error':
                        continue
                    elif response.json()["error"]["message"] == \
                            'Exceeded the daily limit of uploaded images for your account':
                        print('Переключение на второй аккаунт')

                        self.active_token = self.token[1]
                        self.active_secret_key = self.secret_key[1]

                        files = {
                            'image': open(img, 'rb'),
                            'secret_key': (None, self.active_secret_key),
                        }
                        response = requests.post('https://api.imageban.ru/v1', headers=headers, files=files)
                        if response.json()['status'] == 200:
                            img_short_link.append(f"[URL=https://imageban.ru][IMG]{response.json()['data']['link']}"
                                                  f"[/IMG][/URL]")
                        else:
                            print(f'Не удалось загрузить {img}')
                    continue
                except FileNotFoundError:
                    continue
                self.article_save_imgs[img_url] = img_short_link

    def write_final_file_data1(self):
        try:
            if not os.path.isdir('./final_data/'):
                os.mkdir('./final_data/')
            columns = ['N', 'O', 'P']
            wb = load_workbook(filename=self.read_data1_file)
            ws = wb.active

            ws['N8'] = 'Ссылки на фотографии'
            date_now = datetime.datetime.now()
            for article in self.article_save_imgs:
                for i, link in enumerate(self.article_save_imgs[article]):
                    for row in ws.iter_cols(min_col=2, max_col=2, min_row=9):
                        for cell in row:
                            if cell.value.strip().split(' ')[0][3:] in article:
                                ws[f'{columns[i]}{cell.row}'] = link

            file_name = f'./final_data/data1_final_{date_now.strftime("%d-%m-%y_%H-%M")}.xlsx'
            wb.save(filename=file_name)
            print(f'Файл {file_name} сохранён')
        except Exception as exc:
            print(f'Ошибка {exc} в записи итогового файла')
            with open('error.txt', 'a', encoding='utf-8') as file:
                file.write(f'{datetime.datetime.now().strftime("%d-%m-%y %H:%M")} '
                           f'Ошибка {exc} в записи итогового файла, функция - write_final_file_data1()\n')

    def write_final_file_data2(self):
        print('there write data2')
        try:
            if not os.path.isdir('./final_data/'):
                os.mkdir('./final_data/')
            columns = ['Q', 'R', 'S']
            wb = load_workbook(filename=self.read_data2_file)
            ws = wb.active

            ws['Q9'] = 'Ссылки на фотографии'
            date_now = datetime.datetime.now()
            for article in self.article_save_imgs:
                for i, link in enumerate(self.article_save_imgs[article]):
                    for row in ws.iter_cols(min_col=8, max_col=8, min_row=10):
                        for cell in row:
                            print(cell.value)
                            print(cell.value[:-3])
                            if cell.value.strip()[:-3] in article:
                                ws[f'{columns[i]}{cell.row}'] = link

            file_name = f'./final_data/data2_final_{date_now.strftime("%d-%m-%y_%H-%M")}.xlsx'
            wb.save(filename=file_name)
            print(f'Файл {file_name} сохранён')
        except Exception as exc:
            print(f'Ошибка {exc} в записи итогового файла')
            with open('error.txt', 'a', encoding='utf-8') as file:
                file.write(f'{datetime.datetime.now().strftime("%d-%m-%y %H:%M")} '
                           f'Ошибка {exc} в записи итогового файла, функция - write_final_file_data2()\n')

    def run(self):
        # try:
        print('Начало работы')
        self.open_token_file()
        # self.read_file()
        # print('Получаю артикул товаров и ссылки на них')
        # # self.get_article_number()
        # # print('\rАртикулы получил')
        # print('---------------------------\n')
        # print('Получаю ссылки на товары')
        # self.get_link_prodicts()
        # print('\nСсылки получены')
        # print('---------------------------\n')
        # print('Ищу изображения товаров')
        # self.get_link_img()
        # print('\nИзображения получены')
        # print('---------------------------\n')
        # print('Скачиваю изображения')
        # asyncio.run(self.save_images_run_async())
        # print('\nСкачивание завершено')
        # print('---------------------------\n')
        # print('Измененяю размер изображений')
        # self.resize_img()
        # print('\rРазмеры изменены')
        # print('---------------------------\n')
        # print('Загружаю изображения на фотохостинг')
        # self.sending_to_fotohosting()
        # print('\nЗагрузка завершена')
        # print('---------------------------\n')
        # print('Записываю в итоговый файл data1_final')
        # self.write_final_file_data1()
        print('Записываю в итоговый файл data2_final')
        self.write_final_file_data2()
        print('Работа завершена')
        # print('Для выхода нажмите Enter')
        # input()
        # shutil.rmtree('./img/')
        # print('---------------------------\n')
        # except Exception as exc:
        #     print(f'Произошла ошибка {exc}')
        #     print('Для выхода нажмите Enter')
        #     input()
        #     print('---------------------------\n')


# def main():
#     p = Parser()
#     p.run()
#
# if __name__ == '__main__':
#     main()


ll = ['2102']
print(str(*ll))